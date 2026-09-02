#!/usr/bin/env python3
"""
TutorPay 강사관리 구글시트(xlsx 사본) → 정규화 JSON 추출 스크립트

사용법:
  python3 scripts/extract_from_sheet.py <시트.xlsx> [출력.json]
  (기본 출력: data/tutorpay-seed.json)

- 요구사항 정리서 5장(데이터 정리·이전 규칙)을 그대로 구현한다.
- 단가·세전·세후를 4장 규칙으로 재계산해 시트 캐시값과 대조하고, 불일치 건수를 출력한다.
- 원본 값은 임의로 채우지 않는다(공란은 null 유지, 경고 목록으로 출력).
"""
import json
import math
import re
import sys
from collections import Counter, OrderedDict
from datetime import datetime, date

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("openpyxl 이 필요합니다: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------- 상수
GRADES = [
    ("S등급", "리더 강사", 1),
    ("A등급", "일반 강사", 2),
    ("B등급", "인턴강사", 3),
    ("아코연구원", "내부 인력", 4),
]
GRADE_CODES = {g[0] for g in GRADES}

# 단가표 열(시트 순서) → (payType, role)
RATE_COLUMNS = [
    ("관내", "주강사"), ("관외", "주강사"), ("아코센터", "주강사"),
    ("관내", "보조강사"), ("관외", "보조강사"), ("아코센터", "보조강사"),
    ("기관지급", ""), ("주(주말교육)", ""), ("교구정리", ""),
]
PAY_TYPES = ["관내", "관외", "아코센터", "기관지급", "주(주말교육)", "교구정리", "수동기입"]
ROLES = ["주강사", "보조강사"]
INSTRUCTOR_REGION_PREFIXES = ["원주", "강릉", "춘천", "충청", "철원", "태백", "동해"]

# 5.1 콘텐츠 표준명 ← 별칭
CONTENT_ALIASES = OrderedDict([
    ("큐브로이드", ["큐브로이드", "큐브"]),
    ("카미봇", ["카미봇", "카미"]),
    ("카미봇 미로", ["카미봇 미로", "카미봇미로"]),
    ("레고 에센셜", ["레고 에센셜", "레고에센셜", "에센셜"]),
    ("레고 프라임", ["레고 프라임", "레고프라임", "프라임"]),
    ("레고 스파이크", ["레고 스파이크", "레고스파이크", "스파이크"]),
    ("레고", ["레고"]),
    ("로비코", ["로비코"]),
    ("핑퐁", ["핑퐁"]),
    ("핑퐁 웜봇", ["핑퐁 웜봇", "핑퐁웜봇", "웜봇"]),
    ("핑퐁 오토카", ["핑퐁 오토카", "핑퐁오토카", "오토카"]),
    ("핑퐁 모노", ["핑퐁 모노", "핑퐁모노"]),
    ("핑퐁 듀오", ["핑퐁 듀오", "핑퐁듀오"]),
    ("모디", ["모디"]),
    ("아티보", ["아티보"]),
    ("로보독", ["로보독"]),
    ("꼭두", ["꼭두"]),
    ("큐로", ["큐로"]),
    ("앤트봇", ["앤트봇"]),
    ("블록봇", ["블록봇"]),
    ("유갓로봇", ["유갓로봇"]),
    ("드론", ["드론"]),
    ("VR", ["VR", "vr"]),
    ("AI 투닝", ["AI 투닝", "투닝 AI", "AI투닝", "투닝AI", "투닝"]),
    ("제미나이", ["제미나이", "Gemini", "gemini", "제미니"]),
    ("미리캔버스(창업)", ["미리캔버스_창업", "미리캔버스(창업)", "미리캔버스 창업"]),
    ("창업특강", ["창업특강"]),
    ("Golf 특화", ["Golf 특화", "골프 특화", "골프특화"]),
    ("에듀테크", ["에듀테크"]),
    ("프로젝트", ["프로젝트"]),
    ("어린이날부스", ["어린이날부스", "어린이날 부스"]),
    ("페스티벌", ["페스티벌"]),
])
ALIAS_TO_STD = {}
for std, aliases in CONTENT_ALIASES.items():
    for a in aliases:
        ALIAS_TO_STD[a.replace(" ", "").lower()] = std


# ---------------------------------------------------------------- 유틸
def s(v):
    """셀 값을 공백 제거 문자열로. 공란은 None"""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    t = str(v).strip()
    return t if t != "" else None


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def to_iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)):  # 시리얼 날짜
        from datetime import timedelta
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).date().isoformat()
    t = s(v)
    if not t:
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(t, fmt).date().isoformat()
        except ValueError:
            pass
    return None


TIME_TOKEN = re.compile(r"^(\d{1,3}):(\d{1,2})$")


def parse_time_token(tok):
    """'9:10' → '09:10', '10:3' → '10:30', '017:00' → '17:00'. 실패 시 None"""
    m = TIME_TOKEN.match(tok)
    if not m:
        return None
    hh, mm = m.group(1), m.group(2)
    hh = str(int(hh))  # 앞의 0 제거
    if len(mm) == 1:
        mm = mm + "0"  # 분이 1자리면 뒤에 0 보충
    h, mi = int(hh), int(mm)
    if h > 24 or mi > 59:
        return None
    return f"{h:02d}:{mi:02d}"


def parse_time_range(v):
    """5.3 규칙. 반환: (start, end, ok, raw)"""
    raw = s(v)
    if raw is None:
        return None, None, True, None
    t = re.sub(r"\s+", "", raw)
    t = re.sub(r"[∼～\-–—]", "~", t)
    parts = t.split("~")
    if len(parts) != 2:
        return None, None, False, raw
    st, en = parse_time_token(parts[0]), parse_time_token(parts[1])
    if st is None or en is None:
        return None, None, False, raw
    return st, en, True, raw


def classify_institution(name):
    """4.5 규칙: 키워드 포함 여부, 순서대로"""
    if name is None:
        return None
    for kw, typ in (("초등학교", "초등"), ("중학교", "중등"), ("고등학교", "고등"),
                    ("유치원", "유치원"), ("어린이집", "어린이집")):
        if kw in name:
            return typ
    return "기타 기관"


def institution_region(name):
    m = re.search(r"_([가-힣]+)$", name)
    return m.group(1) if m else None


def instructor_region_prefix(name):
    for p in INSTRUCTOR_REGION_PREFIXES:
        if name.startswith(p):
            return p
    return None


def normalize_content(raw):
    """5.1 규칙. 반환 (normalized or None, tokens(std list), unknown tokens)"""
    r = s(raw)
    if r is None:
        return None, [], []
    tokens = [t.strip() for t in re.split(r"[/,]", r) if t.strip()]
    stds, unknown = [], []
    for t in tokens:
        key = t.replace(" ", "").lower()
        std = ALIAS_TO_STD.get(key)
        if std is None:
            unknown.append(t)
            std = t  # 알 수 없는 표기는 그대로 둔다(경고)
        if std not in stds:
            stds.append(std)
    return " / ".join(stds), stds, unknown


def unit_price(rate, grade, pay_type, role, manual_price):
    """4.3 단가 결정 규칙 (시트 I열 수식과 동일한 결과)"""
    if not pay_type:
        return 0
    if pay_type == "수동기입":
        return int(manual_price or 0)
    if grade not in rate:
        return 0
    g = rate[grade]
    if pay_type in ("관내", "관외", "아코센터"):
        r = "보조강사" if role == "보조강사" else "주강사"
        return g[(pay_type, r)]
    if pay_type in ("기관지급", "주(주말교육)", "교구정리"):
        return g[(pay_type, "")]
    return g[("관내", "주강사")]  # SWITCH 기본값


def net_amount(gross):
    """4.4 세후 = floor(세전 × 0.967), 정수 오차 방지"""
    return math.floor(gross * 967 / 1000)


# ---------------------------------------------------------------- 메인
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "data/tutorpay-seed.json"

    wb = openpyxl.load_workbook(src, data_only=True)
    warnings = []

    # ---- 등급별 단가표
    ws = wb["등급별 단가표"]
    rate = {}
    rate_items = []
    for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
        code = s(row[0])
        if code not in GRADE_CODES:
            continue
        rate[code] = {}
        for i, (pt, role) in enumerate(RATE_COLUMNS):
            amt = int(num(row[i + 1]) or 0)
            rate[code][(pt, role)] = amt
            rate_items.append({"grade": code, "payType": pt, "role": role, "amount": amt})
    missing = GRADE_CODES - set(rate)
    if missing:
        warnings.append({"type": "rate_missing_grade", "detail": sorted(missing)})

    # ---- 강사 정보
    ws = wb["강사 정보"]
    instructors = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        name = s(row[0])
        if not name:
            continue
        if name in seen:
            warnings.append({"type": "instructor_duplicate", "row": i, "name": name})
            continue
        seen.add(name)
        grade = s(row[1])
        if grade is not None and grade not in GRADE_CODES:
            warnings.append({"type": "instructor_unknown_grade", "row": i, "name": name, "grade": grade})
            grade = None
        note = s(row[4])
        instructors.append({
            "name": name,
            "grade": grade,                       # None = 미등록
            "phone": s(row[2]),
            "region": s(row[3]) or instructor_region_prefix(name),
            "isActive": not (note and note.lower() == "out"),
            "note": note,
        })
    inst_by_name = {x["name"]: x for x in instructors}

    # ---- 기준목록 (기관)
    ws = wb["기준목록"]
    institutions = []
    seen_inst = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = s(row[3])
        if not name:
            continue
        if name in seen_inst:
            warnings.append({"type": "institution_duplicate_in_list", "name": name})
            continue
        seen_inst.add(name)
        institutions.append({
            "name": name,
            "type": classify_institution(name),
            "region": institution_region(name),
            "isActive": True,
        })

    # ---- 강의배정기록
    ws = wb["강의배정기록"]
    lectures = []
    mismatches = []
    content_counter = Counter()
    unknown_content = Counter()
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        name = s(row[1])
        d = to_iso_date(row[0])
        if not name:
            if d:
                warnings.append({"type": "row_without_instructor", "row": i, "date": d})
            continue
        if not d:
            warnings.append({"type": "row_without_date", "row": i, "instructor": name})
            continue
        if name not in inst_by_name:
            warnings.append({"type": "unknown_instructor", "row": i, "name": name})
            inst_by_name[name] = {"name": name, "grade": None, "phone": None,
                                  "region": instructor_region_prefix(name), "isActive": True, "note": "시트 강사 정보에 없음"}
            instructors.append(inst_by_name[name])

        start, end, ok, raw_time = parse_time_range(row[2])
        if not ok:
            warnings.append({"type": "time_unparsed", "row": i, "raw": raw_time})
        institution = s(row[3])
        if institution and institution not in seen_inst:
            warnings.append({"type": "institution_not_in_list", "row": i, "name": institution})
            seen_inst.add(institution)
            institutions.append({"name": institution, "type": classify_institution(institution),
                                 "region": institution_region(institution), "isActive": True})
        content_raw = s(row[4])
        content, tokens, unknown = normalize_content(content_raw)
        for t in tokens:
            content_counter[t] += 1
        for u in unknown:
            unknown_content[u] += 1
        sessions = num(row[5])
        role = s(row[6]) or "주강사"
        if role not in ROLES:
            warnings.append({"type": "unknown_role", "row": i, "role": role})
        pay_type = s(row[7])
        if pay_type and pay_type not in PAY_TYPES:
            warnings.append({"type": "unknown_pay_type", "row": i, "payType": pay_type})
        sheet_unit = num(row[8])
        sheet_gross = num(row[9])
        sheet_net = num(row[10])
        is_paid = bool(row[11]) if row[11] is not None else False
        is_done = (s(row[12]) == "완료")
        headcount = num(row[13])
        manual = num(row[16])
        note = s(row[17])

        grade = inst_by_name[name]["grade"]
        calc_unit = unit_price(rate, grade, pay_type, role, manual)
        calc_gross = int(round((sessions or 0) * calc_unit))
        calc_net = net_amount(calc_gross)

        # 시트 캐시값과 대조 (수식이 계산된 값이 있을 때만)
        if sheet_unit is not None and int(sheet_unit) != calc_unit:
            mismatches.append({"row": i, "field": "unitPrice", "sheet": sheet_unit, "calc": calc_unit})
        if sheet_gross is not None and int(sheet_gross) != calc_gross:
            mismatches.append({"row": i, "field": "gross", "sheet": sheet_gross, "calc": calc_gross})
        if sheet_net is not None and int(sheet_net) != calc_net:
            mismatches.append({"row": i, "field": "net", "sheet": sheet_net, "calc": calc_net})

        if not pay_type:
            warnings.append({"type": "pay_type_blank", "row": i, "date": d, "instructor": name, "institution": institution})
        if sessions is None:
            warnings.append({"type": "sessions_blank", "row": i, "date": d, "instructor": name, "institution": institution})
        if pay_type == "수동기입" and not manual:
            warnings.append({"type": "manual_price_missing", "row": i, "date": d, "instructor": name})

        lectures.append({
            "sheetRow": i,
            "date": d,
            "startTime": start,
            "endTime": end,
            "timeRaw": raw_time if not ok else None,
            "instructor": name,
            "institution": institution,
            "content": content,
            "contentRaw": content_raw if content_raw != content else None,
            "sessions": sessions,
            "role": role,
            "payType": pay_type,
            # 스냅샷: 시트가 계산해 둔 값이 있으면 그대로, 없으면 재계산값
            "unitPrice": int(sheet_unit) if sheet_unit is not None else calc_unit,
            "grossAmount": int(sheet_gross) if sheet_gross is not None else calc_gross,
            "netAmount": int(sheet_net) if sheet_net is not None else calc_net,
            "manualPrice": int(manual) if manual is not None else None,
            "isPaid": is_paid,
            "isDone": is_done,
            "headcount": int(headcount) if headcount is not None else None,
            "note": note,
        })

    # 콘텐츠 마스터: 표준명 목록(사용된 것 + 정의된 것)
    contents = []
    for std, aliases in CONTENT_ALIASES.items():
        contents.append({"name": std, "aliases": [a for a in aliases if a != std], "usage": content_counter.get(std, 0)})
    for u, c in unknown_content.items():
        warnings.append({"type": "content_unknown_alias", "alias": u, "count": c})
        if u not in {x["name"] for x in contents}:
            contents.append({"name": u, "aliases": [], "usage": c, "unverified": True})

    lectures.sort(key=lambda x: (x["date"], x["startTime"] or "", x["institution"] or "", x["role"] != "주강사", x["instructor"]))

    result = {
        "meta": {
            "source": src.split("/")[-1],
            "extractedAt": datetime.now().isoformat(timespec="seconds"),
            "counts": {
                "instructors": len(instructors),
                "institutions": len(institutions),
                "contents": len(contents),
                "lectures": len(lectures),
                "mismatches": len(mismatches),
                "warnings": len(warnings),
            },
        },
        "grades": [{"code": c, "label": l, "sort": o} for c, l, o in GRADES],
        "rateTable": {"effectiveFrom": "2026-01-01", "memo": "2026년 1월 기준 (구글시트 등급별 단가표)", "items": rate_items},
        "instructors": instructors,
        "institutions": institutions,
        "contents": contents,
        "lectures": lectures,
        "warnings": warnings,
        "mismatches": mismatches,
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)

    # ---- 리포트
    print(f"[extract] lectures={len(lectures)} instructors={len(instructors)} institutions={len(institutions)} contents={len(contents)}")
    print(f"[verify ] mismatch {len(mismatches)}")
    for m in mismatches[:20]:
        print("   ", m)
    wc = Counter(w["type"] for w in warnings)
    print(f"[warn   ] {dict(wc)}")
    for w in warnings:
        if w["type"] in ("pay_type_blank", "sessions_blank", "manual_price_missing", "time_unparsed", "content_unknown_alias", "unknown_instructor", "institution_not_in_list"):
            print("   ", w)
    # 월별 합계 (정산리포트 대조용)
    monthly = {}
    for l in lectures:
        ym = l["date"][:7]
        m = monthly.setdefault(ym, {"count": 0, "sessions": 0.0, "gross": 0, "net": 0})
        m["count"] += 1
        m["sessions"] += l["sessions"] or 0
        m["gross"] += l["grossAmount"]
        m["net"] += l["netAmount"]
    for ym in sorted(monthly):
        m = monthly[ym]
        print(f"[month  ] {ym}: {m['count']}건 {m['sessions']:g}차시 세전 {m['gross']:,} 세후 {m['net']:,}")
    print(f"[out    ] {out}")


if __name__ == "__main__":
    main()
